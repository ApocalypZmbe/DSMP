"""
Analizador ILD1302 — Medición Automática (Arduino autónomo)
===========================================================
Lee el CSV exportado por el Setup Tool de Micro-Epsilon.

El Arduino (sketch ILD1302_auto_30deg.ino) mueve el plato en
modo AUTÓNOMO con esta secuencia:
  1. Espera 8 s inicial (sincronización del sensor).
  2. Mueve PASO_ANGULO° y se queda 3 s quieto en cada posición.
  3. Repite hasta completar 360° × REPETICIONES × sentidos.

Este script detecta todos los períodos estables (≥ TIEMPO_QUIETO_MIN s)
y los asigna secuencialmente a cada posición de la secuencia.

Flujo de medición:
  1. Inicia la grabación en el Setup Tool.
  2. Enciende / resetea el Arduino — sin tocar el monitor serie.
  3. El Arduino espera 8 s, luego gira solo (3 s por posición).
  4. Al terminar, detén la grabación y carga el CSV aquí.

Uso:
  - Doble clic en lanzar.bat                  (recomendado)
  - python analizador_automatico.py           (abre diálogo)
  - python analizador_automatico.py dato.csv  (ruta directa)

By Maurcio Alejandro Diaz Barrera
https://github.com/ApocalypZmbe
"""

# ═══════════════════════════════════════════════════════════════
#  VERIFICACIÓN E INSTALACIÓN AUTOMÁTICA DE LIBRERÍAS
# ═══════════════════════════════════════════════════════════════
import sys
import subprocess

_REQUERIDAS = ["numpy", "pandas", "openpyxl", "matplotlib"]
_faltantes = []
for _p in _REQUERIDAS:
    try:
        __import__(_p)
    except ImportError:
        _faltantes.append(_p)

if _faltantes:
    print("=" * 55)
    print("  Instalando librerías faltantes (solo la primera vez)...")
    print("=" * 55)
    for _p in _faltantes:
        print(f"  → {_p}")
    print()
    try:
        subprocess.check_call([sys.executable, "-m", "pip", "install", *_faltantes])
        print("\n  Instalación completada.\n")
    except subprocess.CalledProcessError:
        print("\n  [ERROR] No se pudieron instalar las librerías.")
        print("  Ejecuta manualmente en la terminal:")
        print(f"    pip install {' '.join(_faltantes)}")
        input("\n  Presiona Enter para salir...")
        sys.exit(1)

# ═══════════════════════════════════════════════════════════════
#  IMPORTACIONES
# ═══════════════════════════════════════════════════════════════
import numpy as np
import pandas as pd
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt
import matplotlib.ticker as ticker

# ═══════════════════════════════════════════════════════════════
#  ┌─────────────────────────────────────────────────────────┐
#  │                    CONFIGURACIÓN                        │
#  │          (modifica solo este bloque)                    │
#  └─────────────────────────────────────────────────────────┘
# ═══════════════════════════════════════════════════════════════

# ── Geometría del plato ──────────────────────────────────────
PASO_ANGULO          = 30          # ° por posición  (debe coincidir con el .ino)
ANGULO_MAX           = 360         # ° total del recorrido
REPETICIONES         = 5           # ciclos por sentido (debe coincidir con CICLOS del .ino)
SENTIDO_PRIMERO      = "Horario"   # "Horario"  o  "Antihorario"
MEDIR_AMBOS_SENTIDOS = True        # False = solo mide SENTIDO_PRIMERO
NUEVO_CERO_ENTRE_SENTIDOS = False  # True = toma nueva pausa de referencia en 0° antes del sentido 2

# ── Detección de períodos estables ──────────────────────────
UMBRAL_STD           = 0.15        # mm  — variación máxima permitida para "quieto"
VENTANA_STD_SEG      = 0.20        # s   — ventana del rolling std
TIEMPO_QUIETO_MIN    = 2.50        # s   — mínimo para detectar pausa (3s del Arduino con margen)
TIEMPO_MARGEN        = 0.30        # s   — margen descartado al inicio y fin de cada período

# ── Referencia inicial (posición 0°) ─────────────────────────
TIEMPO_PAUSA_INICIO  = 7.00        # s   — pausa inicial del Arduino (8s, con margen)

# ── Archivo de salida ────────────────────────────────────────
NOMBRE_SALIDA        = "resultados_automatico.xlsx"

# ═══════════════════════════════════════════════════════════════
#  (no modificar debajo de esta línea)
# ═══════════════════════════════════════════════════════════════

_SENTIDO_2 = "Antihorario" if SENTIDO_PRIMERO == "Horario" else "Horario"


def seleccionar_archivo():
    if len(sys.argv) > 1:
        return sys.argv[1]
    try:
        import tkinter as tk
        from tkinter import filedialog
        root = tk.Tk()
        root.withdraw()
        root.attributes("-topmost", True)
        ruta = filedialog.askopenfilename(
            title="Seleccionar CSV del sensor ILD1302",
            filetypes=[("Archivos CSV", "*.csv"), ("Todos los archivos", "*.*")]
        )
        root.destroy()
        if not ruta:
            print("No se seleccionó archivo. Saliendo.")
            sys.exit(0)
        return ruta
    except Exception as e:
        print(f"No se pudo abrir el diálogo de archivo: {e}")
        print("Uso: python analizador_manual.py ruta_al_archivo.csv")
        sys.exit(1)


ARCHIVO_CSV = seleccionar_archivo()

# ── Colores Excel ─────────────────────────────────────────────
AZ_OSC = "1F3864"; AZ_MED = "2E75B6"; AZ_CLR = "BDD7EE"
RJ     = "C0392B"; RJ_CLR = "FADBD8"
GR     = "F2F2F2"; BL     = "FFFFFF"
AM     = "FFC000"; NA_CLR = "FDEBD0"

def _fill(h):   return PatternFill("solid", fgColor=h)
def _font(bold=False, color="000000", size=11, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Arial")
def _align(h="center", v="center"):
    return Alignment(horizontal=h, vertical=v, wrap_text=False)
def _border():
    s = Side(style="thin")
    return Border(left=s, right=s, top=s, bottom=s)

def cel(ws, row, col, value=None, fill=None, bold=False, color="000000",
        size=11, italic=False, align_h="center", num_fmt=None, border=True):
    c = ws.cell(row, col, value)
    if fill:     c.fill = _fill(fill)
    c.font       = _font(bold=bold, color=color, size=size, italic=italic)
    c.alignment  = _align(h=align_h)
    if border:   c.border = _border()
    if num_fmt:  c.number_format = num_fmt
    return c


# ═══════════════════════════════════════════════════════════════
#  LECTURA DEL CSV
# ═══════════════════════════════════════════════════════════════
def _parse_timestamps(series):
    # Formato datetime completo: "2026-05-08 10:11:55.000376"
    dt = pd.to_datetime(series, errors="coerce")
    if dt.notna().sum() > len(series) // 2:
        return dt.astype("int64") / 1e9  # nanosegundos → segundos

    # Fallback: formato legado "MM:SS.sss"
    def _legacy(s):
        try:
            p = str(s).strip().split(":")
            return float(p[0]) * 60 + float(p[1])
        except:
            return np.nan
    return series.apply(_legacy)


def cargar_csv(ruta):
    ruta = Path(ruta)
    if not ruta.exists():
        raise FileNotFoundError(f"No se encontró: {ruta}")
    print(f"Cargando: {ruta.name}")
    df = pd.read_csv(ruta, skiprows=3, header=None, usecols=[0, 1],
                     encoding="latin1", on_bad_lines="skip")
    df.columns = ["ts", "dist"]
    df["dist"] = pd.to_numeric(df["dist"], errors="coerce")
    df = df.dropna(subset=["dist"]).reset_index(drop=True)

    df["t"] = _parse_timestamps(df["ts"])
    df = df.dropna(subset=["t"]).reset_index(drop=True)

    if df.empty:
        raise ValueError(
            "No se pudieron leer los timestamps del archivo.\n"
            f"   Ejemplo de valor en columna timestamp: {df['ts'].iloc[0] if not df.empty else 'N/A'}"
        )

    df["t"] -= df["t"].iloc[0]
    dur = df["t"].iloc[-1]
    print(f"   {len(df):,} filas | {dur:.1f} s | ~{len(df)/dur:.0f} Hz")
    return df


# ═══════════════════════════════════════════════════════════════
#  DETECCIÓN DE PERÍODOS QUIETOS
# ═══════════════════════════════════════════════════════════════
def detectar_quietos(df):
    t = df["t"].values
    d = df["dist"].values
    diffs = np.diff(t)
    diffs = diffs[diffs > 0]
    dt  = np.median(diffs) if len(diffs) else 0.001
    win = max(int(VENTANA_STD_SEG / dt), 3)
    std_r  = pd.Series(d).rolling(win, center=True, min_periods=2).std().fillna(999).values
    quieto = std_r < UMBRAL_STD

    periodos = []
    en_p = False; i0 = 0
    for i in range(len(quieto)):
        if quieto[i] and not en_p:
            en_p = True; i0 = i
        elif not quieto[i] and en_p:
            en_p = False
            dur = t[i - 1] - t[i0]
            if dur >= TIEMPO_QUIETO_MIN:
                periodos.append((t[i0], t[i - 1], dur, i0, i - 1))
    if en_p:
        dur = t[-1] - t[i0]
        if dur >= TIEMPO_QUIETO_MIN:
            periodos.append((t[i0], t[-1], dur, i0, len(t) - 1))

    result = []
    for (ti, tf, dur, ii, fi) in periodos:
        mask = (t >= ti + TIEMPO_MARGEN) & (t <= tf - TIEMPO_MARGEN)
        vals = d[mask]
        vals = vals[~np.isnan(vals)]
        result.append({"t_ini": ti, "t_fin": tf, "dur": dur, "vals": vals})

    print(f"   Períodos quietos (≥{TIEMPO_QUIETO_MIN}s): {len(result)}")
    return result


# ═══════════════════════════════════════════════════════════════
#  SECUENCIA Y ASIGNACIÓN
# ═══════════════════════════════════════════════════════════════
def construir_secuencia():
    angulos = list(range(PASO_ANGULO, ANGULO_MAX + 1, PASO_ANGULO))
    seq = []

    if MEDIR_AMBOS_SENTIDOS and NUEVO_CERO_ENTRE_SENTIDOS:
        # Todos los ciclos del sentido 1 primero, luego todos los del sentido 2
        for ciclo in range(1, REPETICIONES + 1):
            for ang in angulos:
                seq.append({"sentido": SENTIDO_PRIMERO, "angulo": ang, "ciclo": ciclo})
        for ciclo in range(1, REPETICIONES + 1):
            for ang in angulos:
                seq.append({"sentido": _SENTIDO_2, "angulo": ang, "ciclo": ciclo})
    else:
        for ciclo in range(1, REPETICIONES + 1):
            for ang in angulos:
                seq.append({"sentido": SENTIDO_PRIMERO, "angulo": ang, "ciclo": ciclo})
            if MEDIR_AMBOS_SENTIDOS:
                for ang in angulos:
                    seq.append({"sentido": _SENTIDO_2, "angulo": ang, "ciclo": ciclo})
    return seq


def asignar(periodos, secuencia):
    idx_inicio = None
    dist_ref   = np.nan

    if TIEMPO_PAUSA_INICIO > 0:
        for i, p in enumerate(periodos):
            if p["dur"] >= TIEMPO_PAUSA_INICIO:
                idx_inicio = i
                dist_ref   = np.mean(p["vals"]) if len(p["vals"]) > 0 else np.nan
                print(f"   Pausa inicio : {p['t_ini']:.1f}s → {p['t_fin']:.1f}s  ({p['dur']:.1f}s)")
                ref_str = f"{dist_ref:.4f} mm" if not np.isnan(dist_ref) else "N/A"
                print(f"   Referencia 0°: {ref_str}")
                break

    if idx_inicio is None:
        print("   Sin pausa de inicio detectada. Usando todos los períodos como mediciones.")
        mediciones = periodos
    else:
        mediciones = periodos[idx_inicio + 1:]

    # Cuántas mediciones pertenecen al sentido 1 (para saber dónde insertar el nuevo cero)
    n_s1 = sum(1 for s in secuencia if s["sentido"] == SENTIDO_PRIMERO)
    n_esperados = len(secuencia) + (1 if NUEVO_CERO_ENTRE_SENTIDOS and MEDIR_AMBOS_SENTIDOS else 0)
    print(f"   Períodos de medición: {len(mediciones)} | Esperados: {n_esperados}")
    if len(mediciones) < n_esperados:
        print(f"   Faltan {n_esperados - len(mediciones)} períodos.")

    rows = []
    med_idx   = 0
    ref_actual = dist_ref   # puede cambiar si NUEVO_CERO_ENTRE_SENTIDOS=True

    for i, slot in enumerate(secuencia):
        # Consumir nueva pausa de referencia justo antes del sentido 2
        if (NUEVO_CERO_ENTRE_SENTIDOS and MEDIR_AMBOS_SENTIDOS
                and i == n_s1 and med_idx < len(mediciones)):
            p_ref2 = mediciones[med_idx]
            med_idx += 1
            ref_actual = np.mean(p_ref2["vals"]) if len(p_ref2["vals"]) > 0 else ref_actual
            ref_str2 = f"{ref_actual:.4f} mm" if not np.isnan(ref_actual) else "N/A"
            print(f"   Nuevo punto 0° (entre sentidos): {ref_str2}  "
                  f"({p_ref2['t_ini']:.1f}s - {p_ref2['t_fin']:.1f}s, {p_ref2['dur']:.1f}s)")

        if med_idx < len(mediciones):
            p    = mediciones[med_idx]; med_idx += 1
            vals = p["vals"]
            n    = len(vals)
            prom = np.mean(vals)          if n > 0 else np.nan
            std  = np.std(vals, ddof=1)   if n > 1 else (0. if n == 1 else np.nan)
            err  = std / np.sqrt(n)       if n > 0 else np.nan
            error = (prom - ref_actual
                     if not np.isnan(prom) and not np.isnan(ref_actual)
                     else np.nan)
            rows.append({**slot, "n": n, "prom": prom, "std": std,
                         "err_std": err, "error": error, "dist_ref": ref_actual})
        else:
            rows.append({**slot, "n": 0, "prom": np.nan, "std": np.nan,
                         "err_std": np.nan, "error": np.nan, "dist_ref": ref_actual})
    return rows, dist_ref


# ═══════════════════════════════════════════════════════════════
#  CÁLCULOS
# ═══════════════════════════════════════════════════════════════
def calcular_calibracion(rows, sentido):
    angulos = list(range(PASO_ANGULO, ANGULO_MAX + 1, PASO_ANGULO))
    datos   = [r for r in rows if r["sentido"] == sentido]
    result  = []
    for ang in angulos:
        errores = [r["error"] for r in datos
                   if r["angulo"] == ang and not np.isnan(r["error"])]
        n = len(errores)
        if n == 0:
            result.append({"angulo": ang, "error_prom": np.nan, "std_error": np.nan, "n": 0})
            continue
        ep  = np.mean(errores)
        std = np.std(errores, ddof=1) if n > 1 else 0.
        result.append({"angulo": ang, "error_prom": ep, "std_error": std, "n": n})
    return result


def calcular_reposicionamiento(rows, dist_ref):
    angulos = list(range(PASO_ANGULO, ANGULO_MAX + 1, PASO_ANGULO))
    hor  = [r for r in rows if r["sentido"] == "Horario"]
    anti = [r for r in rows if r["sentido"] == "Antihorario"]
    result = [{"angulo": 0, "prom_h": dist_ref, "prom_a": dist_ref,
               "diferencia": 0.0, "std_dif": 0.0, "n": 0}]
    for ang in angulos:
        vals_h = [r["prom"] for r in hor  if r["angulo"] == ang and not np.isnan(r["prom"])]
        vals_a = [r["prom"] for r in anti if r["angulo"] == ang and not np.isnan(r["prom"])]
        n = min(len(vals_h), len(vals_a))
        if n == 0:
            result.append({"angulo": ang, "prom_h": np.nan, "prom_a": np.nan,
                           "diferencia": np.nan, "std_dif": np.nan, "n": 0})
            continue
        difs = [abs(h - a) for h, a in zip(vals_h[:n], vals_a[:n])]
        result.append({"angulo": ang,
                        "prom_h": np.mean(vals_h), "prom_a": np.mean(vals_a),
                        "diferencia": np.mean(difs),
                        "std_dif": np.std(difs, ddof=1) if n > 1 else 0.0, "n": n})
    return result


# ═══════════════════════════════════════════════════════════════
#  GRÁFICA
# ═══════════════════════════════════════════════════════════════
def generar_grafica(calib_s1, calib_s2=None):
    plt.rcParams.update({"font.family": "DejaVu Sans",
                          "axes.spines.top": False, "axes.spines.right": False,
                          "grid.alpha": 0.35, "figure.dpi": 140})
    fig, ax = plt.subplots(figsize=(11, 5.5), facecolor="white")
    ax.set_facecolor("white")

    def _plot(calib, color, marker, label):
        ang = [0] + [c["angulo"] for c in calib]
        std = [0] + [c["std_error"] if not np.isnan(c["std_error"]) else 0 for c in calib]
        ax.plot(ang, std, f"{marker}-", color=color, lw=2, ms=7,
                markerfacecolor="white", markeredgewidth=2, label=label)

    flecha1 = "↻" if SENTIDO_PRIMERO == "Horario" else "↺"
    color1  = "#1a6faf" if SENTIDO_PRIMERO == "Horario" else "#c0392b"
    _plot(calib_s1, color1, "o", f"{flecha1} Sentido {SENTIDO_PRIMERO}")

    if calib_s2 is not None:
        flecha2 = "↺" if SENTIDO_PRIMERO == "Horario" else "↻"
        color2  = "#c0392b" if SENTIDO_PRIMERO == "Horario" else "#1a6faf"
        _plot(calib_s2, color2, "s", f"{flecha2} Sentido {_SENTIDO_2}")

    ax.axhline(0, color="gray", lw=0.8, linestyle="--", alpha=0.6)
    ax.set_title(
        f"Calibración — Desviación Estándar vs Ángulo"
        f"  (mín {TIEMPO_QUIETO_MIN:.0f} s/posición · paso {PASO_ANGULO}°)",
        fontsize=13, fontweight="bold", pad=12
    )
    ax.set_xlabel("Grados (°)", fontsize=11)
    ax.set_ylabel("Desviación estándar (mm)", fontsize=11)
    ax.xaxis.set_major_locator(ticker.MultipleLocator(PASO_ANGULO))
    ax.grid(True, linestyle="--")
    ax.legend(fontsize=10, framealpha=0.8)

    ruta = str(Path(__file__).parent / "grafica_automatica.png")
    fig.savefig(ruta, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print("   Gráfica guardada")
    return ruta


# ═══════════════════════════════════════════════════════════════
#  EXCEL — HOJA MEDICIONES
# ═══════════════════════════════════════════════════════════════
def hoja_mediciones(wb, rows, dist_ref):
    ws = wb.active
    ws.title = "Mediciones"
    ws.sheet_view.showGridLines = False
    angulos = list(range(PASO_ANGULO, ANGULO_MAX + 1, PASO_ANGULO))
    n_cols  = len(angulos) + 2
    col_fin = get_column_letter(n_cols)

    ws.merge_cells(f"A1:{col_fin}1")
    c = ws.cell(1, 1,
        f"MEDICIONES — SENSOR ILD1302  |  Paso {PASO_ANGULO}°  |  "
        f"Máx {ANGULO_MAX}°  |  {REPETICIONES} ciclos/sentido")
    c.fill = _fill(AZ_OSC); c.font = _font(bold=True, color="FFFFFF", size=14)
    c.alignment = _align(); ws.row_dimensions[1].height = 30

    ref_str = f"{dist_ref:.4f} mm" if not np.isnan(dist_ref) else "N/A"
    ws.merge_cells(f"A2:{col_fin}2")
    c = ws.cell(2, 1,
        f"Umbral std: {UMBRAL_STD} mm  |  Quietud mín: {TIEMPO_QUIETO_MIN} s  |  "
        f"Margen: {TIEMPO_MARGEN} s  |  Referencia 0°: {ref_str}")
    c.fill = _fill(AZ_MED); c.font = _font(italic=True, color="FFFFFF", size=10)
    c.alignment = _align(); ws.row_dimensions[2].height = 16

    sentidos = [SENTIDO_PRIMERO] + ([_SENTIDO_2] if MEDIR_AMBOS_SENTIDOS else [])

    def bloque(sentido, fila):
        color_h   = AZ_MED if sentido == "Horario" else RJ
        color_clr = AZ_CLR if sentido == "Horario" else RJ_CLR
        flecha    = "↻"    if sentido == "Horario" else "↺"
        datos     = [r for r in rows if r["sentido"] == sentido]

        ws.merge_cells(f"A{fila}:{col_fin}{fila}")
        c = ws.cell(fila, 1, f"  {flecha}  SENTIDO {sentido.upper()}")
        c.fill = _fill(color_h); c.font = _font(bold=True, color="FFFFFF", size=12)
        c.alignment = _align(h="left"); ws.row_dimensions[fila].height = 22; fila += 1

        cel(ws, fila, 1, "Ciclo",       fill=AZ_OSC, bold=True, color="FFFFFF", size=10)
        cel(ws, fila, 2, "0° Ref (mm)", fill=AZ_OSC, bold=True, color="FFFFFF", size=10)
        ws.column_dimensions["A"].width = 8
        ws.column_dimensions["B"].width = 13
        for ci, ang in enumerate(angulos, 3):
            cel(ws, fila, ci, f"{ang}°  (mm)", fill=color_h, bold=True, color="FFFFFF", size=10)
            ws.column_dimensions[get_column_letter(ci)].width = 13
        ws.row_dimensions[fila].height = 20; fila += 1

        for ciclo in range(1, REPETICIONES + 1):
            fill_bg = color_clr if ciclo % 2 == 0 else BL
            cel(ws, fila, 1, ciclo, fill=color_h, bold=True, color="FFFFFF")
            # Usar la referencia almacenada en la fila (puede diferir por sentido)
            _rep = [r for r in datos if r["ciclo"] == ciclo]
            _ref = _rep[0]["dist_ref"] if _rep else dist_ref
            ref_val = round(_ref, 4) if not np.isnan(_ref) else "N/A"
            cel(ws, fila, 2, ref_val,
                fill=AZ_OSC if ciclo == 1 else GR,
                color="FFFFFF" if ciclo == 1 else "666666",
                num_fmt="0.0000", italic=ciclo > 1)
            for ci, ang in enumerate(angulos, 3):
                rep = [r for r in datos if r["angulo"] == ang and r["ciclo"] == ciclo]
                if rep:
                    v = rep[0]["prom"]
                    if np.isnan(v):
                        cel(ws, fila, ci, "N/D", fill=fill_bg, italic=True, color="999999")
                    else:
                        cel(ws, fila, ci, round(v, 4), fill=fill_bg, num_fmt="0.0000")
                else:
                    cel(ws, fila, ci, "N/D", fill=fill_bg, italic=True, color="999999")
            ws.row_dimensions[fila].height = 17; fila += 1
        return fila

    fila = 4
    for s in sentidos:
        fila = bloque(s, fila)
        fila += 1


# ═══════════════════════════════════════════════════════════════
#  EXCEL — HOJA RESULTADOS
# ═══════════════════════════════════════════════════════════════
def hoja_resultados(wb, rows, calib_s1, dist_ref, ruta_png, calib_s2=None):
    ws = wb.create_sheet("Resultados")
    ws.sheet_view.showGridLines = False
    angulos = list(range(PASO_ANGULO, ANGULO_MAX + 1, PASO_ANGULO))
    max_cols = max(len(angulos) + 2, 8)
    col_fin  = get_column_letter(max_cols)

    ws.merge_cells(f"A1:{col_fin}1")
    modo = "Ambos sentidos" if MEDIR_AMBOS_SENTIDOS else SENTIDO_PRIMERO
    c = ws.cell(1, 1,
        f"CALIBRACIÓN — SENSOR ILD1302  |  Paso {PASO_ANGULO}°  |  {modo}")
    c.fill = _fill(AZ_OSC); c.font = _font(bold=True, color="FFFFFF", size=14)
    c.alignment = _align(); ws.row_dimensions[1].height = 30

    ref_str = f"{dist_ref:.4f} mm" if not np.isnan(dist_ref) else "N/A"
    ws.merge_cells(f"A2:{col_fin}2")
    c = ws.cell(2, 1,
        f"Referencia 0°: {ref_str}  |  Error = Distancia medida − Referencia")
    c.fill = _fill(AZ_MED); c.font = _font(italic=True, color="FFFFFF", size=10)
    c.alignment = _align(); ws.row_dimensions[2].height = 16
    fila = 4

    # Tabla 1: Reposicionamiento (solo si se midieron ambos sentidos)
    if MEDIR_AMBOS_SENTIDOS and calib_s2 is not None:
        repos   = calcular_reposicionamiento(rows, dist_ref)
        difs_v  = [r["diferencia"] for r in repos
                   if not np.isnan(r["diferencia"]) and r["angulo"] != 0]
        std_v   = [r["std_dif"]    for r in repos
                   if not np.isnan(r["std_dif"])    and r["angulo"] != 0]
        ep_rep  = np.mean(difs_v) if difs_v else np.nan
        std_rep = np.mean(std_v)  if std_v  else np.nan

        ws.merge_cells(f"A{fila}:C{fila}")
        c = ws.cell(fila, 1, "Tabla 1 — Reposicionamiento")
        c.fill = _fill(AZ_OSC); c.font = _font(bold=True, color="FFFFFF", size=12)
        c.alignment = _align(h="left"); ws.row_dimensions[fila].height = 22; fila += 1

        for ci, h in enumerate(["Giro", "Error Promedio (mm)", "Desv. estándar (mm)"], 1):
            cel(ws, fila, ci, h, fill=AZ_MED, bold=True, color="FFFFFF", size=10)
            ws.column_dimensions[get_column_letter(ci)].width = 30
        ws.row_dimensions[fila].height = 22; fila += 1

        f1 = "↻" if SENTIDO_PRIMERO == "Horario" else "↺"
        f2 = "↺" if SENTIDO_PRIMERO == "Horario" else "↻"
        for label, clr in [
            (f"{f1} {SENTIDO_PRIMERO} / {f2} {_SENTIDO_2}", AZ_CLR),
            (f"{f2} {_SENTIDO_2} / {f1} {SENTIDO_PRIMERO}", RJ_CLR),
        ]:
            cel(ws, fila, 1, label, fill=clr, align_h="left")
            cel(ws, fila, 2,
                round(ep_rep, 4)  if not np.isnan(ep_rep)  else "N/D",
                fill=clr, num_fmt="0.0000")
            cel(ws, fila, 3,
                round(std_rep, 4) if not np.isnan(std_rep) else "N/D",
                fill=clr, num_fmt="0.0000")
            ws.row_dimensions[fila].height = 17; fila += 1
        fila += 1

    def tabla_calib(sentido, calib, fila, num):
        color_h   = AZ_MED if sentido == "Horario" else RJ
        color_clr = AZ_CLR if sentido == "Horario" else RJ_CLR
        flecha    = "↻"    if sentido == "Horario" else "↺"
        col_fin_t = get_column_letter(len(calib) + 2)

        ws.merge_cells(f"A{fila}:{col_fin_t}{fila}")
        c = ws.cell(fila, 1, f"Tabla {num} — Calibración {flecha} {sentido}")
        c.fill = _fill(color_h); c.font = _font(bold=True, color="FFFFFF", size=12)
        c.alignment = _align(h="left"); ws.row_dimensions[fila].height = 22; fila += 1

        for ci, h in enumerate(
            ["Valor nominal (°)", "Error Promedio (mm)", "Desv. Std del error (mm)"], 1
        ):
            cel(ws, fila, ci, h, fill=AZ_OSC, bold=True, color="FFFFFF", size=10)
            ws.column_dimensions[get_column_letter(ci)].width = 26
        ws.column_dimensions["A"].width = 18
        ws.row_dimensions[fila].height = 22; fila += 1

        for idx, row in enumerate(calib):
            fill_bg = color_clr if idx % 2 == 0 else BL
            cel(ws, fila, 1, f"{row['angulo']}°", fill=color_h, bold=True, color="FFFFFF")
            ep = row["error_prom"]; sd = row["std_error"]
            cel(ws, fila, 2,
                round(ep, 4) if not np.isnan(ep) else "N/D",
                fill=AM, bold=True, num_fmt="0.0000")
            cel(ws, fila, 3,
                round(sd, 4) if not np.isnan(sd) else "N/D",
                fill=NA_CLR, num_fmt="0.0000")
            ws.row_dimensions[fila].height = 17; fila += 1
        return fila

    num1 = 1 if not MEDIR_AMBOS_SENTIDOS else (2 if SENTIDO_PRIMERO == "Horario" else 3)
    fila = tabla_calib(SENTIDO_PRIMERO, calib_s1, fila, num1)

    if MEDIR_AMBOS_SENTIDOS and calib_s2 is not None:
        fila += 1
        num2 = 3 if SENTIDO_PRIMERO == "Horario" else 2
        fila = tabla_calib(_SENTIDO_2, calib_s2, fila, num2)

    fila += 2
    ws.merge_cells(f"A{fila}:{col_fin}{fila}")
    c = ws.cell(fila, 1, "Gráfica — Desviación Estándar vs Ángulo")
    c.fill = _fill(AZ_OSC); c.font = _font(bold=True, color="FFFFFF", size=12)
    c.alignment = _align(); ws.row_dimensions[fila].height = 22; fila += 1

    img = XLImage(ruta_png); img.width = 750; img.height = 400
    img.anchor = f"A{fila}"; ws.add_image(img)
    for ci in range(1, 5):
        ws.column_dimensions[get_column_letter(ci)].width = 30


# ═══════════════════════════════════════════════════════════════
#  MAIN
# ═══════════════════════════════════════════════════════════════
def main():
    n_pos     = ANGULO_MAX // PASO_ANGULO
    n_sentidos = 2 if MEDIR_AMBOS_SENTIDOS else 1

    print("=" * 60)
    print("  Analizador ILD1302 — Modo Automático (Arduino)")
    print(f"  Paso: {PASO_ANGULO}°  |  Máx: {ANGULO_MAX}°  |  "
          f"{n_pos} posiciones  |  {REPETICIONES} ciclos")
    print(f"  Sentido primero: {SENTIDO_PRIMERO}  |  "
          f"Sentidos: {n_sentidos}  |  "
          f"Quietud mín: {TIEMPO_QUIETO_MIN} s")
    print("=" * 60)

    df = cargar_csv(ARCHIVO_CSV)

    print("\nDetectando períodos quietos...")
    per = detectar_quietos(df)

    print("\nConstruyendo secuencia...")
    seq = construir_secuencia()
    print(f"   {n_pos} pos × {REPETICIONES} ciclos × {n_sentidos} sentido(s) "
          f"= {len(seq)} mediciones esperadas")

    rows, dist_ref = asignar(per, seq)
    print(f"   Slots con datos: {sum(1 for r in rows if r['n'] > 0)}/{len(seq)}")

    print("\nCalculando tablas...")
    calib_s1 = calcular_calibracion(rows, SENTIDO_PRIMERO)
    calib_s2 = calcular_calibracion(rows, _SENTIDO_2) if MEDIR_AMBOS_SENTIDOS else None

    print("\nGenerando gráfica...")
    ruta_png = generar_grafica(calib_s1, calib_s2)

    print("\nConstruyendo Excel...")
    wb = Workbook()
    hoja_mediciones(wb, rows, dist_ref)
    hoja_resultados(wb, rows, calib_s1, dist_ref, ruta_png, calib_s2)

    salida = Path(__file__).parent / NOMBRE_SALIDA
    wb.save(str(salida))
    print(f"   Guardado: {salida}")
    print("\nListo.\n")
    input("Presiona Enter para cerrar...")


if __name__ == "__main__":
    main()
